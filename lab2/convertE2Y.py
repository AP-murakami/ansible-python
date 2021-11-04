# import 宣言
import openpyxl
import yaml
import logging
import sys
from rich.logging import RichHandler

# ストリームハンドラの設定とファイルハンドラの設定
rich_handler: RichHandler = RichHandler(rich_tracebacks=True)
rich_handler.setFormatter(logging.Formatter("%(asctime)s:%(lineno)d:%(levelname)s:%(message)s"))

  # ファイルハンドラの設定
logfile = logging.FileHandler('converte2y.log')
logfile.setFormatter(logging.Formatter("%(asctime)s:%(lineno)d:%(levelname)s:%(message)s"))
  
# ログレベルの設定とハンドラの関連づけ
logging.basicConfig(level=logging.DEBUG, handlers=[rich_handler,logfile])

# Logger名の指定
logger = logging.getLogger(__name__)

# main 
def main():
  # Excel ファイルを指定する
  wbook = openpyxl.load_workbook("AnsiblePythonLab.xlsx")
  #logger.debug()
  
  # Excel ファイルのシートリストを取得する
  sheets = wbook.sheetnames
  logger.debug("SheetList :"+ str(sheets))
  
  # sheet の枚数分繰り返す
  for sheetname in sheets:
    
    logger.info("-- Processing : " + sheetname + " sheet --")
    sheet = wbook[sheetname] # 処理するシートの取得
    hostname = sheet["C6"]
    ip = sheet["C7"]
    cider = sheet["C8"]
    gateway = sheet["C9"]
    dns = sheet["C10"]
    packages = list()
  
    # インストールパッケージリストの取得
    packages = getList(sheet,row_no=15,col_no=2,max_row_no=27) 
    # 起動/有効化サービスリストの取得
    services = getList(sheet,row_no=15,col_no=5,max_row_no=27) 
    # セキュリティファイアウォール(インバウンド許可リスト)の取得
    firewalls = getList(sheet,row_no=30,col_no=3,max_row_no=35)
    # 追加ユーザリストの取得
    users = getList(sheet,row_no=40,col_no=2,max_row_no=47)
     # 追加ユーザリストのコメント取得
    users_comment = getList(sheet,row_no=40,col_no=3,max_row_no=47)
    
    # yaml 出力データの整形
    out_data = {'hostanme': hostname.value,
                'ip': ip.value,
                'cider': cider.value,
                'gateway': gateway.value,
                'dns': dns.value,
                'packages': packages,
                'services': services,
                'firewalls': firewalls,
                'users': users,
                'users_comment': users_comment
               }

    # 出力ファイルの指定
    yaml_output = open("host_vars/"+ sheetname + ".yaml", 'w') 

    # 各シートの情報を yaml に変換して出力
    yaml_output.write(yaml.dump(out_data))
    yaml_output.close()

    # yaml出力内容の出力
    logger.debug("hostname: " + str(hostname.value))
    logger.debug("ip : " + str(ip.value))
    logger.debug("cider: " + str(cider.value))
    logger.debug("gateway: " + str(gateway.value))
    logger.debug("dns: " + str(dns.value))
    logger.debug("packages: " + str(packages))
    logger.debug("services: " + str(services))
    logger.debug("firewalls: " + str(firewalls))
    logger.debug("users: " + str(users))
    logger.debug("users_comment: " + str(users_comment))

  # man() 正常終了
  return 0

# Excel シートからリストで取得
def getList(sheet,row_no,col_no,max_row_no):
  convlist = list()

  # 起動/有効化サービスリストの処理
  while row_no < max_row_no:
      cell_value = sheet.cell(row=row_no,column=col_no).value
      if cell_value == None:
          row_no = row_no + 1
          continue
      convlist.append(cell_value)
      row_no = row_no + 1
  #logger.info("convlist: " + str(convlist))
  return convlist

if __name__ == '__main__':
    logger.info('START MAIN PROCESS')
    main() 